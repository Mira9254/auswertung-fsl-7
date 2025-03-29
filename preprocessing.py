from openpyxl import load_workbook, Workbook

participant_name_column = "G"


def extract_participant_responses(sheet, row_number):
    """
    Extrahiert die Antworten eines Teilnehmers aus der Excel-Tabelle.
    """

    participant_name = sheet[f"{participant_name_column}{row_number}"].value

    # Startspalte für die Antworten (Frage 2 beginnt bei Spalte T)
    start_column = ord("Q")
    # Anzahl der Fragen
    total_questions = 28
    # Abstand der Fragen
    answer_offset = 6
    # Anzahl der Antwortmöglichkeiten (A, B, C)
    number_of_answer_options = 3

    # Hier werden die Antworten des Teilnehmers gespeichert
    answers = []

    # Gehe durch alle 28 Fragen
    for question in range(1, total_questions + 1):
        # Berechne alle Spalten für die Antwortmöglichkeiten
        # Liste für die Antwortspalten
        answer_columns = []
        # Schleife über die Anzahl der Antwortmöglichkeiten (A, B, C)
        for i in range(number_of_answer_options):
            # Berechne den Index der Antwortspalte
            answer_column_index = start_column + (question - 1) * answer_offset + i
            # Berechne den ersten Buchstaben der Spalte
            first_letter = chr(ord("A") + (answer_column_index - ord("A")) // 26 - 1)
            # Berechne den zweiten Buchstaben der Spalte
            second_letter = chr(ord("A") + (answer_column_index - ord("A")) % 26)
            # Wenn der erste Buchstabe kleiner als "A" ist
            if first_letter < "A":
                # Füge nur den zweiten Buchstaben zur Liste hinzu
                answer_columns.append(second_letter)
            else:
                # Füge beide Buchstaben zur Liste hinzu
                answer_columns.append(first_letter + second_letter)

        # Lese die Werte der Antwortspalten aus
        answer_selections = [sheet[f"{answer_column}{row_number}"].value for answer_column in answer_columns]

        # Prüfen, welche Antwort gewählt wurde (1 steht für die Auswahl)
        if answer_selections[0] == 1:
            answers.append("A")
        elif answer_selections[1] == 1:
            answers.append("B")
        elif answer_selections[2] == 1:
            answers.append("C")
        else:
            # Für den Fall, dass keine gültige Antwort vorliegt
            answers.append(None)

    # Rückgabe des Teilnehmernamens und der Antworten
    return participant_name, answers


def process_excel_data(input_file):
    """
    Extrahiert die Antworten der Schüler aus der Excel-Datei für die weitere Verarbeitung.
    """

    # Excel-Datei mit den Daten laden
    workbook = load_workbook(input_file)
    sheet = workbook.active

    # Neue Excel-Datei erstellen
    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = "Results"

    # Kopfzeile für die Ergebnisse
    output_sheet.append(["Teilnehmer", *[f"Frage {i+1}" for i in range(28)]])

    # Verarbeitung aller Teilnehmer
    # Startzeile für die Teilnehmerdaten
    start_row = 3
    current_row = start_row
    # Schleife über alle Teilnehmerdaten
    while True:
        # Teilnehmername aus der aktuellen Zeile
        participant_name = sheet[f"{participant_name_column}{current_row}"].value

        # Schleife nach letztem Teilnehmer beenden
        if not participant_name:
            break

        # Antworten auswerten
        _, answers = extract_participant_responses(sheet, current_row)

        # Ergebnisse speichern
        output_sheet.append([participant_name, *answers])

        # Zur nächsten Zeile wechseln
        current_row += 1

    # Excel-Datei schließen
    workbook.close()
    # Rückgabe der verarbeiteten Daten
    return output_workbook
