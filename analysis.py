from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

evaluation_matrix = [
    # 1. Seite
    ["s", "ui", "e"],
    ["e", "s", "ui"],
    ["ui", "e", "s"],
    ["s", "ui", "e"],
    ["e", "s", "ui"],
    ["ui", "e", "s"],
    ["s", "ui", "e"],
    # 2. Seite
    ["ui", "e", "s"],
    ["e", "s", "ui"],
    ["s", "ui", "e"],
    ["e", "ui", "s"],
    ["ui", "e", "s"],
    ["s", "e", "ui"],
    ["e", "ui", "s"],
    # 3. Seite
    ["e", "s", "ui"],
    ["s", "e", "ui"],
    ["e", "ui", "s"],
    ["ui", "s", "e"],
    ["ui", "s", "e"],
    ["ui", "e", "s"],
    ["s", "e", "ui"],
    # 4. Seite
    ["ui", "s", "e"],
    ["s", "e", "ui"],
    ["ui", "s", "e"],
    ["e", "ui", "s"],
    ["s", "e", "ui"],
    ["s", "ui", "e"],
    ["e", "ui", "s"],
]

facetten = [
    "Fähigkeit zur Einschätzung des eigenen Lernstands",
    "Fähigkeit adäquate Lernziele zu setzen",
    "Wahl einer geeigneten Lernstrategie",
    "Anwendungsgüte der Lernstrategie",
    "Fähigkeit zur Feststellung des eigenen Lernfortschritts",
    "Fähigkeit zur Anpassung des eigenen Lernens",
    "Überprüfung und Feststellung des Lernergebnisses",
]

ordered_types = [
    "Selbstreguliert",
    "Überwiegend selbstreguliert",
    "Ansatzweise selbstreguliert",
    "Mischtyp selbstreguliert / external reguliert",
    "Mischtyp selbstreguliert / unreflektiert-impulsiv",
    "Überwiegend external reguliert",
    "Überwiegend unreflektiert-impulsiv",
    "Ansatzweise external reguliert",
    "Ansatzweise unreflektiert-impulsiv",
    "External reguliert",
    "Unreflektiert-impulsiv",
    "Mischtyp external reguliert / unreflektiert-impulsiv",
    "Keine Zuordnung",
]


def get_facette(facette, row):
    """
    Berechnet die Facetten-Ergebnisse für eine gegebene Facette.
    """

    result = {"s": 0, "ui": 0, "e": 0}

    for i in range(facette, len(row), 7):
        antwortABC = row[i].value
        if antwortABC:
            antwortNumber = ord(antwortABC) - ord("A")
            antwort = evaluation_matrix[i - 1][antwortNumber]
            result[antwort] += 1

    return result


def get_final_result(facette):
    if facette["s"] == 4:
        return "Selbstreguliert"
    if facette["s"] == 3:
        return "Überwiegend selbstreguliert"

    if facette["s"] == 2:
        if facette["e"] == 1 and facette["ui"] == 1:
            return "Ansatzweise selbstreguliert"
        if facette["e"] == 2:
            return "Mischtyp selbstreguliert / external reguliert"
        if facette["ui"] == 2:
            return "Mischtyp selbstreguliert / unreflektiert-impulsiv"

    if facette["s"] == 1:
        if facette["e"] == 3:
            return "Überwiegend external reguliert"
        if facette["ui"] == 3:
            return "Überwiegend unreflektiert-impulsiv"
        if facette["e"] == 2 and facette["ui"] == 1:
            return "Ansatzweise external reguliert"
        if facette["e"] == 1 and facette["ui"] == 2:
            return "Ansatzweise unreflektiert-impulsiv"

    if facette["s"] == 0:
        if facette["e"] == 4:
            return "External reguliert"
        if facette["ui"] == 4:
            return "Unreflektiert-impulsiv"
        if facette["e"] == 2 and facette["ui"] == 2:
            return "Mischtyp external reguliert / unreflektiert-impulsiv"
        if facette["e"] == 3 and facette["ui"] == 1:
            return "Überwiegend external reguliert"
        if facette["e"] == 1 and facette["ui"] == 3:
            return "Überwiegend unreflektiert-impulsiv"

    return f"Keine Zuordnung für diesen Fall ({facette})"


def generate_evaluation_report(workbook):
    """
    Erstellt einen Bericht über die Auswertung der Excel-Daten.
    """

    sheet = workbook.active
    result_workbook = Workbook()

    # Klassenübersicht vorbereiten
    class_summary = {i: {t: 0 for t in ordered_types} for i in range(7)}

    # Teilnehmer verarbeiten
    students = []
    for row_index, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        student_name = row[0].value
        if student_name:
            students.append((student_name, row))

    # Teilnehmer nach Namen sortieren
    students.sort(key=lambda x: x[0])

    # Ergebnisse für jeden Teilnehmer sammeln
    for student_name, row in students:
        # Teilnehmer-Sheet erstellen
        result_sheet = result_workbook.create_sheet(title=student_name[:30])

        # Überschriften
        result_sheet.append(["Facette", "Details", "Bewertung"])
        # Überschriften hervorheben
        for cell in result_sheet[1]:
            cell.font = Font(bold=True)

        # Facetten-Ergebnisse sammeln und Klassenübersicht aktualisieren
        for facette_index in range(7):
            facette_result = get_facette(facette_index + 1, row)
            final_result = get_final_result(facette_result)

            # Für Klassenübersicht zählen
            if final_result in class_summary[facette_index]:
                class_summary[facette_index][final_result] += 1
            else:
                class_summary[facette_index]["Keine Zuordnung"] += 1

            # Teilnehmer-Sheet schreiben
            result_sheet.append([facetten[facette_index], str(facette_result), final_result])

        # Spaltenbreiten anpassen
        result_sheet.column_dimensions["A"].width = 50
        result_sheet.column_dimensions["B"].width = 20
        result_sheet.column_dimensions["C"].width = 50

    # Klassenübersicht erstellen
    class_sheet = result_workbook.create_sheet(title="Klassenübersicht", index=0)
    headers = ["Facette"] + ordered_types
    class_sheet.append(headers)

    # Überschriften hervorheben
    for cell in class_sheet[1]:
        cell.font = Font(bold=True)

    # Daten einfügen
    for facette_index in range(7):
        row_data = [facetten[facette_index]]
        for t in ordered_types:
            row_data.append(class_summary[facette_index][t])
        class_sheet.append(row_data)

    # Summenzeile hinzufügen
    sum_row = ["Summe"]
    for col_idx in range(2, len(ordered_types) + 2):
        col_letter = get_column_letter(col_idx)
        sum_formula = f"=SUM({col_letter}2:{col_letter}{class_sheet.max_row})"
        sum_row.append(sum_formula)
    class_sheet.append(sum_row)

    # Spaltenbreiten anpassen
    class_sheet.column_dimensions["A"].width = 50
    for col_idx in range(2, len(ordered_types) + 2):
        col_letter = get_column_letter(col_idx)
        class_sheet.column_dimensions[col_letter].width = 30

    # Leeres Standard-Sheet entfernen
    if "Sheet" in result_workbook.sheetnames:
        del result_workbook["Sheet"]

    return result_workbook
